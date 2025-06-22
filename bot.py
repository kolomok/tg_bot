from telegram import Update, ReplyKeyboardMarkup, InputFile
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters
from openpyxl import Workbook, load_workbook
import os
from dotenv import load_dotenv

# 📁 Папка для хранения Excel-файлов пользователей
USER_DATA_DIR = "user_excels"
os.makedirs(USER_DATA_DIR, exist_ok=True)

# 🔑 Получаем токен из переменных окружения
TOKEN = os.environ.get("BOT_TOKEN")

# 📁 Генерация пути к Excel-файлу пользователя
def get_excel_path(user_id: int) -> str:
     desktop = os.path.join(os.path.expanduser("~"), "Desktop")
     os.makedirs(desktop, exist_ok=True)
     filename = f"{user_id}.xlsx"
     return os.path.join(desktop, filename)

# 📄 Создание Excel-файла, если его нет
def init_excel(file_path: str):
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Товар", "Количество", "Цена/штука", "Цена/общая"])
        wb.save(file_path)

async def export_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    file_path = get_excel_path(user_id)

    if os.path.exists(file_path):
        with open(file_path, "rb") as f:
            await update.message.reply_document(document=InputFile(f, filename=os.path.basename(file_path)))
    else:
        await update.message.reply_text("Файл ещё не создан. Сначала добавьте данные.")

# 📋 Клавиатура
reply_keyboard = [["Добавить данные", "📤 Экспортировать файл"]]
markup = ReplyKeyboardMarkup(reply_keyboard, resize_keyboard=True)
# 🔘 Команда /start
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Привет! Я бот для записи данных в Excel.\n\nНажми кнопку ниже, чтобы добавить данные:",
        reply_markup=markup
    )

# ➕ Запрос ввода данных
async def prompt_for_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Введите данные в формате:\n`Товар, Количество, Цена/штука, Цена/общая`",
        parse_mode="Markdown"
    )
    context.user_data["awaiting_data"] = True

# 📥 Обработка текстовых сообщений
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if context.user_data.get("awaiting_data"):
        user_id = update.message.from_user.id
        file_path = get_excel_path(user_id)
        init_excel(file_path)

        try:
            parts = [p.strip() for p in update.message.text.split(",")]
            if len(parts) != 4:
                raise ValueError("Ожидается 4 значения, разделённые запятыми.")

            wb = load_workbook(file_path)
            ws = wb.active
            ws.append(parts)
            wb.save(file_path)

            await update.message.reply_text("✅ Данные успешно сохранены.")
        except Exception as e:
            await update.message.reply_text(f"❌ Ошибка: {e}")
        finally:
            context.user_data["awaiting_data"] = False
    else:
        await update.message.reply_text("Нажмите 'Добавить данные', чтобы внести запись.")

# 🚀 Запуск бота
if __name__ == "__main__":
    load_dotenv()

    TOKEN = os.environ.get("BOT_TOKEN")
    RENDER_HOSTNAME = os.environ.get("RENDER_EXTERNAL_HOSTNAME")
    PORT = int(os.environ.get("PORT", 8443))

    if not TOKEN or not RENDER_HOSTNAME:
        raise ValueError("BOT_TOKEN или RENDER_EXTERNAL_HOSTNAME не установлены.")

    app = ApplicationBuilder().token(TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("export", export_file))
    app.add_handler(MessageHandler(filters.Text(["Добавить данные"]), prompt_for_data))
    app.add_handler(MessageHandler(filters.Text(["📤 Экспортировать файл"]), export_file))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    app.run_webhook(listen="0.0.0.0",
        port=PORT,
        webhook_url=f"https://{RENDER_HOSTNAME}/webhook"
    )

